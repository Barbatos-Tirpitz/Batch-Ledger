from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ACCENT = "2E5F55"
INK = "211E19"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color=ACCENT, end_color=ACCENT)
note_font = Font(name="Arial", italic=True, color="5B564A", size=10)
body_font = Font(name="Arial", size=11)
title_font = Font(name="Arial", bold=True, size=16, color=INK)
subtitle_font = Font(name="Arial", bold=True, size=13, color=ACCENT)
thin = Side(style="thin", color="DEDACB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

def style_body_rows(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

# ---------------- Instructions sheet ----------------
ws = wb.active
ws.title = "Start Here"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 90

ws["A1"] = "Batch Ledger — Import Templates"
ws["A1"].font = title_font
ws["A2"] = "Fill in each sheet, then export it as CSV (File > Save As > CSV) to import into the app."
ws["A2"].font = note_font

rows = [
    ("", ""),
    ("How this works", "header"),
    ("This workbook has one tab per import type. Each tab already has the exact column headers the app expects, plus a couple of example rows so you can see the format — delete the example rows before adding your real data.", "body"),
    ("", ""),
    ("1. Ingredients tab", "header"),
    ("Your raw materials. Columns: name, unit, cost (cost is optional). The unit column has a dropdown so you don't mistype it.", "body"),
    ("", ""),
    ("2. Products tab", "header"),
    ("Recipes. One row per ingredient used in a product. Columns: product, ingredient, qty_per_unit.", "body"),
    ("", ""),
    ("3. Production Log tab", "header"),
    ("What you actually produced. Columns: date (YYYY-MM-DD), product, qty_produced.", "body"),
    ("", ""),
    ("Import order matters a little", "header"),
    ("Import Ingredients first, then Products, then the Production Log.", "body"),
    ("", ""),
    ("Exporting each tab as its own CSV", "header"),
    ("Right-click the sheet tab > Move or Copy > (new workbook) > Save As > CSV. Repeat once per tab — the app imports one CSV per data type, not the whole workbook.", "body"),
]

r = 4
for text, kind in rows:
    cell = ws.cell(row=r, column=1, value=text)
    if kind == "header":
        cell.font = subtitle_font
    elif kind == "body":
        cell.font = body_font
        ws.row_dimensions[r].height = 30
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ---------------- Ingredients sheet ----------------
ws = wb.create_sheet("Ingredients")
ws.sheet_view.showGridLines = False
headers = ["name", "unit", "cost"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header_row(ws, 1, len(headers))

sample = [
    ["Cocoa butter", "kg", 8.5],
    ["Sugar", "kg", 1.2],
    ["Milk powder", "kg", 4.75],
    ["Vanilla extract", "ml", 0.35],
]
for ridx, row in enumerate(sample, 2):
    for cidx, val in enumerate(row, 1):
        ws.cell(row=ridx, column=cidx, value=val)
style_body_rows(ws, 2, 2 + len(sample) - 1, len(headers))

ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.freeze_panes = "A2"

unit_dv = DataValidation(type="list", formula1='"g,kg,ml,L,pcs"', allow_blank=True, showDropDown=False)
unit_dv.error = "Pick a unit from the list"
unit_dv.prompt = "Choose the unit for this ingredient"
ws.add_data_validation(unit_dv)
unit_dv.add("B2:B500")

note = ws.cell(row=len(sample) + 3, column=1, value="↑ Replace the example rows above with your own ingredients. Cost is optional.")
note.font = note_font

# ---------------- Products sheet ----------------
ws = wb.create_sheet("Products")
ws.sheet_view.showGridLines = False
headers = ["product", "ingredient", "qty_per_unit"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header_row(ws, 1, len(headers))

sample = [
    ["Dark chocolate bar 100g", "Cocoa butter", 0.045],
    ["Dark chocolate bar 100g", "Sugar", 0.02],
    ["Milk chocolate bar 100g", "Cocoa butter", 0.03],
    ["Milk chocolate bar 100g", "Sugar", 0.025],
    ["Milk chocolate bar 100g", "Milk powder", 0.02],
    ["Milk chocolate bar 100g", "Vanilla extract", 0.5],
]
for ridx, row in enumerate(sample, 2):
    for cidx, val in enumerate(row, 1):
        ws.cell(row=ridx, column=cidx, value=val)
style_body_rows(ws, 2, 2 + len(sample) - 1, len(headers))

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 16
ws.freeze_panes = "A2"

note = ws.cell(row=len(sample) + 3, column=1,
               value="↑ One row per ingredient in a recipe — repeat the product name for each of its ingredients.")
note.font = note_font
note2 = ws.cell(row=len(sample) + 4, column=1,
                value="qty_per_unit = how much of that ingredient goes into ONE unit of the product.")
note2.font = note_font

# ---------------- Production Log sheet ----------------
ws = wb.create_sheet("Production Log")
ws.sheet_view.showGridLines = False
headers = ["date", "product", "qty_produced"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header_row(ws, 1, len(headers))

sample = [
    ["2026-07-01", "Dark chocolate bar 100g", 500],
    ["2026-07-01", "Milk chocolate bar 100g", 300],
    ["2026-07-02", "Dark chocolate bar 100g", 420],
]
for ridx, row in enumerate(sample, 2):
    for cidx, val in enumerate(row, 1):
        ws.cell(row=ridx, column=cidx, value=val)
style_body_rows(ws, 2, 2 + len(sample) - 1, len(headers))

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 14
ws.freeze_panes = "A2"

date_dv = DataValidation(type="date", operator="greaterThan", formula1="2000-01-01",
                          allow_blank=True, showErrorMessage=True)
date_dv.error = "Enter a valid date (YYYY-MM-DD)"
date_dv.prompt = "Format: YYYY-MM-DD"
ws.add_data_validation(date_dv)
date_dv.add("A2:A500")
for r in range(2, 2 + len(sample)):
    ws.cell(row=r, column=1).number_format = "YYYY-MM-DD"

note = ws.cell(row=len(sample) + 3, column=1,
               value="↑ product must match a name from the Products tab exactly (case-insensitive).")
note.font = note_font

wb.save("Batch_Ledger_Import_Template.xlsx")
print("saved Batch_Ledger_Import_Template.xlsx")